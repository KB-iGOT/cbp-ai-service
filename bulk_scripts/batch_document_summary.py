"""
Bulk document-summary runner (backend workaround, run in DEV).

Reads a source file (.xlsx or .csv) whose rows carry a state/dept scope with a
"document available = yes" flag, resolves each kept row to the documents in that
scope, and generates their summaries by calling the EXACT same logic the CBP API
uses -- src.api.v1.document_routes._run_document_summary -- so the write-back to
the DB (summary_text / summary_status / summary_error) is identical to a normal
API-triggered summary. Processes in batches of 10.

Source file:
  * .xlsx -- ALL worksheet tabs are read by default (restrict with --sheet).
  * .csv  -- read directly.
  * If a row instead carries a document id (file_id UUID or filename), the runner
    switches to per-document matching automatically.

An optional CSV (state_id, dept_id, state_name, dept_name) can enrich the output
log with names when the source file doesn't already carry them.

USAGE (run in the dev environment, where the DB + GCS bucket are reachable):

  # 1) DRY RUN (default) -- resolves rows to file_ids, prints a plan, writes a
  #    plan CSV. NO Gemini calls, no GCS needed. Verify the detected columns here.
  .venv/bin/python scripts/bulk_summary_runner.py --excel /path/to/source.xlsx --user-id <uuid>

  # 2) Test on a few, for real:
  .venv/bin/python scripts/bulk_summary_runner.py --excel /path/source.xlsx --user-id <uuid> --execute --limit 5

  # 3) Full run:
  .venv/bin/python scripts/bulk_summary_runner.py --excel /path/source.xlsx --user-id <uuid> --execute

Notes:
  * DB target + GCS bucket come from .env (settings.DATABASE_URL, GCP_STORAGE_*).
    Point .env at the dev DB/bucket that actually holds the documents.
  * Already-COMPLETED docs are skipped (like the API). Use --force to regenerate.
  * A stuck IN_PROGRESS doc is auto-recovered when stale; --reset-inprogress forces all.
"""
from __future__ import annotations

import argparse
import asyncio
import contextvars
import csv
import logging
import os
import pkgutil
import re
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Optional

# ── path bootstrap ────────────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from src.core.configs import settings  # noqa: E402  (loads .env)

# Make sure Vertex/GCS credentials are available to child libs (mirrors the app).
if settings.GOOGLE_APPLICATION_CREDENTIALS:
    os.environ.setdefault("GOOGLE_APPLICATION_CREDENTIALS", settings.GOOGLE_APPLICATION_CREDENTIALS)

# Register ALL SQLAlchemy mappers (the server imports every model at startup; a
# standalone script must too, or relationships like 'Role' fail to resolve).
import src.models  # noqa: E402
for _m in pkgutil.iter_modules(src.models.__path__):
    __import__(f"src.models.{_m.name}")

from sqlalchemy import select, and_, or_  # noqa: E402
from src.core.database import sessionmanager  # noqa: E402
from src.models.document import Document  # noqa: E402
from src.crud.document import crud_document  # noqa: E402

# IMPORTANT: import the app's logger module NOW, up front. It runs logging.config.fileConfig(),
# which disables existing loggers and replaces root handlers. If it ran later (it is pulled in
# lazily by `import ...document_routes` inside run_execute), it would silently kill OUR logging
# mid-run -- every line after "EXECUTING" would vanish. Triggering it here (once, cached) means
# our setup_logging() below runs afterwards and its handlers stick.
import src.core.logger  # noqa: E402,F401

log = logging.getLogger("bulk_summary")

# per-task token sink: process_one() sets a fresh dict here; the patched genai call
# accumulates each response's usage_metadata into it (survives asyncio.wait_for's task copy
# because the dict reference is shared).
_usage_sink: contextvars.ContextVar = contextvars.ContextVar("bulk_usage", default=None)

# default log file -- one fresh, timestamped file per run (matches the other bulk_scripts)
RUN_TIMESTAMP = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
DEFAULT_LOG_FILE = str(Path(__file__).resolve().parent / "logs" / f"bulk_summary_runner_{RUN_TIMESTAMP}.log")


def setup_logging(path: str) -> str:
    """Own the logging config: console + file (one fresh file per run). Installed AFTER the
    app's fileConfig has run, and it also re-enables/propagates the app's own 'ai_cbp_service'
    logger so its summary progress lands in the same place."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    root = logging.getLogger()
    for h in list(root.handlers):                    # drop whatever fileConfig installed
        root.removeHandler(h)
    con = logging.StreamHandler()
    con.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    fh = logging.FileHandler(path, mode="w", encoding="utf-8")   # fresh file per run
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s"))
    root.addHandler(con)
    root.addHandler(fh)
    root.setLevel(logging.INFO)
    for name in ("bulk_summary", "ai_cbp_service"):  # re-enable loggers fileConfig may have disabled
        lg = logging.getLogger(name)
        lg.disabled = False
        lg.handlers = []
        lg.propagate = True
        lg.setLevel(logging.INFO)
    logging.getLogger("google").setLevel(logging.WARNING)
    return path

# ══════════════════════════════════════════════════════════════════════════════
# CONFIG -- defaults; every value can be overridden by a CLI flag (see argparse).
# ══════════════════════════════════════════════════════════════════════════════

# Target user (mandatory --user-id CLI arg, parsed in main()): every document lookup
# (scope, file_id, and filename modes) is restricted to documents.uploader_id == user_id,
# so a run only ever touches one user's documents.

EXCEL_PATH = ""          # source file (.xlsx or .csv); set via --excel
EXCEL_SHEET = ""         # xlsx worksheet name; blank => ALL tabs are read
CSV_PATH = ""            # optional state/dept id->name reference CSV; set via --csv
OUT_PATH = ""            # result/plan CSV; blank => auto-named next to the source file

# Column names in the source file. Blank => auto-detect (and printed for you to verify).
FILTER_COL = ""          # the yes/no column to filter on (e.g. "document_available")
FILTER_TRUE_VALUES = {"yes"}   # row kept if its FILTER_COL value (lowercased) is in here
FILE_ID_COL = ""         # column holding the documents.file_id UUID (preferred match key)
FILENAME_COL = ""        # column holding the filename (fallback match key)
STATE_COL = ""           # state_center_id column (used to disambiguate filename matches)
DEPT_COL = ""            # department_id column (used to disambiguate filename matches)

# CSV column names (id->name reference). Blank => auto-detect.
CSV_STATE_ID_COL = ""
CSV_DEPT_ID_COL = ""
CSV_STATE_NAME_COL = ""
CSV_DEPT_NAME_COL = ""

BATCH_SIZE = 10          # max documents summarized concurrently
PER_DOC_TIMEOUT = 1200   # seconds; on timeout the doc is marked FAILED (never left IN_PROGRESS)
RETRIES = 1              # extra in-run attempts if a doc ends FAILED (0 = single attempt)
RESUME_STALE_MINUTES = 30  # on restart, an IN_PROGRESS doc untouched for longer than this is
                           # treated as crashed and reset to NOT_STARTED (auto crash-recovery)

# Candidate header names for auto-detection (normalized: lowercase, alnum only).
_CAND = {
    "filter":    {"documentavailable", "docavailable", "available", "hasdocument",
                  "documentspresent", "document", "documents", "summary", "summaryrequired",
                  "needsummary", "generatesummary", "process", "include", "tosummarize"},
    "file_id":   {"fileid", "documentid", "docid", "uuid", "documentfileid"},
    "filename":  {"filename", "docname", "pdf", "pdfname"},
    "state":     {"statecenterid", "statecenter", "stateid", "state", "orgid",
                  "frameworkid", "statecode"},
    "dept":      {"departmentmdoid", "mdoid", "deptmdoid", "departmentid", "deptid",
                  "department", "dept", "deptcode"},
    "statename": {"statecentername", "statename"},
    "deptname":  {"departmentmdoname", "mdoname", "departmentname", "deptname"},
    "ndocs":     {"numberofdocuments", "numdocuments", "numdocs", "documentcount", "docscount"},
}


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s is not None else ""


def _looks_like_uuid(v) -> bool:
    try:
        uuid.UUID(str(v).strip())
        return True
    except (ValueError, AttributeError, TypeError):
        return False


# ── source loading (.xlsx multi-tab / .csv) ──────────────────────────────────
_SHEET_KEY = "__sheet__"   # reserved per-row keys recording the row's origin
_ROW_KEY = "__row__"


def load_rows(path: str, sheet: str = "") -> tuple[list[str], list[dict]]:
    """Load the source file into (headers, rows). rows are dicts keyed by header, each
    also carrying its origin under _SHEET_KEY/_ROW_KEY.
      * .csv  -> read directly.
      * .xlsx -> read ALL worksheet tabs (or just `sheet` if given). Headers are the
                 union across tabs, so heterogeneous-but-overlapping tabs still work."""
    if path.lower().endswith(".csv"):
        hdrs, rows = load_csv_rows(path)
        for i, r in enumerate(rows, start=2):
            r[_SHEET_KEY], r[_ROW_KEY] = "(csv)", i
        return hdrs, rows

    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = [wb[sheet]] if sheet else list(wb.worksheets)
    all_headers: list[str] = []
    seen: set[str] = set()
    rows: list[dict] = []
    per_sheet: list[tuple] = []
    for ws in sheets:
        it = ws.iter_rows(values_only=True)
        try:
            first = next(it)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(first)]
        for h in headers:
            if h not in seen:
                seen.add(h)
                all_headers.append(h)
        cnt = 0
        for rnum, raw in enumerate(it, start=2):  # header is row 1
            if raw is None or all(c is None for c in raw):
                continue
            row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
            row[_SHEET_KEY], row[_ROW_KEY] = ws.title, rnum
            rows.append(row)
            cnt += 1
        per_sheet.append((ws.title, cnt))
    wb.close()
    if len(per_sheet) > 1:
        log.info(f"read {len(per_sheet)} tabs: " + ", ".join(f"{t}={c}" for t, c in per_sheet))
    return all_headers, rows


def load_csv_rows(path: str) -> tuple[list[str], list[dict]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        return (reader.fieldnames or []), rows


def _origin(r: dict, fallback_idx: int) -> tuple[str, int]:
    """(sheet, row) for a source row, from its reserved keys (fallback to an index)."""
    return r.get(_SHEET_KEY, ""), r.get(_ROW_KEY, fallback_idx)


def detect_col(headers: list[str], override: str, kind: str,
               require_uuid_sample: list | None = None) -> Optional[str]:
    """Pick a column: explicit override wins; else first header whose normalized
    name is a known candidate for `kind`. If require_uuid_sample is given, the
    chosen column's sample values must look like UUIDs."""
    if override:
        match = next((h for h in headers if h == override or _norm(h) == _norm(override)), None)
        if not match:
            raise SystemExit(f"[config] column '{override}' (for {kind}) not found. Headers: {headers}")
        return match
    cands = _CAND.get(kind, set())
    for h in headers:
        if _norm(h) in cands:
            if require_uuid_sample is not None:
                sample = [r.get(h) for r in require_uuid_sample[:20] if r.get(h) is not None]
                if sample and not any(_looks_like_uuid(v) for v in sample):
                    continue
            return h
    return None


# ── DB resolution ────────────────────────────────────────────────────────────
async def _fetch_docs_by_file_ids(user_id: uuid.UUID, ids: list[uuid.UUID]) -> dict[str, Document]:
    out: dict[str, Document] = {}
    async with sessionmanager.session() as db:
        for i in range(0, len(ids), 500):
            chunk = ids[i:i + 500]
            res = await db.execute(select(Document).where(
                Document.file_id.in_(chunk), Document.uploader_id == user_id
            ))
            for d in res.scalars().all():
                out[str(d.file_id)] = d
    return out


async def _fetch_docs_by_filenames(user_id: uuid.UUID, names: list[str]) -> dict[str, list[Document]]:
    out: dict[str, list[Document]] = {}
    async with sessionmanager.session() as db:
        for i in range(0, len(names), 500):
            chunk = names[i:i + 500]
            res = await db.execute(select(Document).where(
                Document.filename.in_(chunk), Document.uploader_id == user_id
            ))
            for d in res.scalars().all():
                out.setdefault(d.filename, []).append(d)
    return out


@dataclass
class Target:
    excel_row: int
    sheet: str = ""                    # source tab (or "(csv)")
    file_id: Optional[str] = None
    filename: Optional[str] = None
    state_id: Optional[str] = None
    dept_id: Optional[str] = None
    doc: Optional[Document] = None
    resolution: str = "pending"        # matched | unresolved | ambiguous
    prior_status: Optional[str] = None
    final_status: Optional[str] = None
    error: Optional[str] = None
    attempts: int = 0
    tok_input: int = 0
    tok_output: int = 0
    tok_thinking: int = 0
    tok_total: int = 0
    names: tuple = field(default=("", ""))  # (state_name, dept_name)


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


async def resolve_targets(rows: list[dict], cols: dict, user_id: uuid.UUID) -> list[Target]:
    """Build Target list from filtered rows and resolve each to a DB Document."""
    targets: list[Target] = []
    for idx, r in enumerate(rows, start=2):
        sheet, rownum = _origin(r, idx)
        targets.append(Target(
            excel_row=rownum,
            sheet=sheet,
            file_id=_s(r.get(cols["file_id"])) if cols["file_id"] else None,
            filename=_s(r.get(cols["filename"])) if cols["filename"] else None,
            state_id=_s(r.get(cols["state"])) if cols["state"] else None,
            dept_id=_s(r.get(cols["dept"])) if cols["dept"] else None,
        ))

    if cols["file_id"]:
        ids = [uuid.UUID(t.file_id) for t in targets if t.file_id and _looks_like_uuid(t.file_id)]
        by_id = await _fetch_docs_by_file_ids(user_id, ids)
        for t in targets:
            if not t.file_id or not _looks_like_uuid(t.file_id):
                t.resolution = "unresolved"
                t.error = "missing/invalid file_id"
                continue
            doc = by_id.get(str(uuid.UUID(t.file_id)))
            if doc:
                t.doc, t.resolution = doc, "matched"
                t.filename = t.filename or doc.filename
            else:
                t.resolution = "unresolved"
                t.error = "file_id not in documents table"
    elif cols["filename"]:
        names = [t.filename for t in targets if t.filename]
        by_name = await _fetch_docs_by_filenames(user_id, names)
        for t in targets:
            if not t.filename:
                t.resolution, t.error = "unresolved", "missing filename"
                continue
            cands = by_name.get(t.filename, [])
            if t.state_id:
                cands = [d for d in cands if str(d.state_center_id) == t.state_id]
            if t.dept_id:
                cands = [d for d in cands if str(d.department_id or "") == t.dept_id]
            if len(cands) == 1:
                t.doc, t.resolution, t.file_id = cands[0], "matched", str(cands[0].file_id)
            elif len(cands) == 0:
                t.resolution, t.error = "unresolved", "filename not found for scope"
            else:
                t.resolution, t.error = "ambiguous", f"{len(cands)} docs match filename+scope"
    else:
        raise SystemExit("[config] neither a file_id column nor a filename column was found in the excel. "
                         "Set --file-id-col or --filename-col.")

    for t in targets:
        if t.doc:
            t.prior_status = t.doc.summary_status
    return targets


async def _fetch_docs_by_scopes(user_id: uuid.UUID, keys: list[tuple]) -> dict[tuple, list[Document]]:
    """keys: list of (state_center_id, department_id|None). Returns {(state,dept): [Document]}.
    Restricted to user_id's documents only."""
    out: dict[tuple, list[Document]] = {}
    async with sessionmanager.session() as db:
        for i in range(0, len(keys), 200):
            conds = []
            for s, d in keys[i:i + 200]:
                conds.append(and_(Document.state_center_id == s,
                                  Document.department_id == d if d else Document.department_id.is_(None)))
            res = await db.execute(select(Document).where(
                Document.uploader_id == user_id, or_(*conds)
            ))
            for doc in res.scalars().all():
                k = (str(doc.state_center_id), str(doc.department_id) if doc.department_id is not None else None)
                out.setdefault(k, []).append(doc)
    return out


async def resolve_scope_targets(rows: list[dict], cols: dict, user_id: uuid.UUID) -> list[Target]:
    """SCOPE mode: the sheet has no per-document id, only (state, dept). Dedupe rows by
    (state_center_id, department_id), then fetch ALL of that scope's documents from the DB
    and emit one Target per document. This is the 'get fileid for each state/dept' flow."""
    scopes: dict[tuple, dict] = {}
    order: list[tuple] = []
    for idx, r in enumerate(rows, start=2):
        sid = _s(r.get(cols["state"])) if cols.get("state") else None
        did = _s(r.get(cols["dept"])) if cols.get("dept") else None
        key = (sid, did)
        if key not in scopes:
            sn = _s(r.get(cols["statename"])) if cols.get("statename") else None
            dn = _s(r.get(cols["deptname"])) if cols.get("deptname") else None
            sheet, rownum = _origin(r, idx)
            scopes[key] = {"row": rownum, "sheet": sheet, "names": (sn or "", dn or ""), "expected": 0}
            order.append(key)
        if cols.get("ndocs"):
            try:
                scopes[key]["expected"] = max(scopes[key]["expected"],
                                              int(float(_s(r.get(cols["ndocs"])) or 0)))
            except (ValueError, TypeError):
                pass

    docs_by_scope = await _fetch_docs_by_scopes(user_id, list(scopes.keys()))
    targets: list[Target] = []
    for key in order:
        sc = scopes[key]
        docs = docs_by_scope.get(key, [])
        if not docs:
            targets.append(Target(excel_row=sc["row"], sheet=sc["sheet"], state_id=key[0], dept_id=key[1],
                                  resolution="unresolved", names=sc["names"],
                                  error="no documents in DB for this state/dept scope"))
            continue
        if sc["expected"] and len(docs) != sc["expected"]:
            log.warning(f"scope ({key[0]}, {key[1]}) source says {sc['expected']} doc(s) "
                        f"but DB has {len(docs)}")
        for d in docs:
            targets.append(Target(excel_row=sc["row"], sheet=sc["sheet"], file_id=str(d.file_id),
                                  filename=d.filename, state_id=key[0], dept_id=key[1], doc=d,
                                  resolution="matched", prior_status=d.summary_status, names=sc["names"]))
    return targets


# ── execution ────────────────────────────────────────────────────────────────
def _install_token_logging() -> None:
    """Wrap google-genai's async generate_content so each summary's token usage is captured
    into the current task's _usage_sink (per-doc), for logging + the report. No-op if already
    patched. _run_document_summary calls this method internally, so we see every summary."""
    try:
        from google.genai.models import AsyncModels
    except Exception as exc:
        log.warning(f"token logging unavailable: {exc}")
        return
    if getattr(AsyncModels, "_bulk_tok_patched", False):
        return
    _orig = AsyncModels.generate_content

    async def _wrapped(self, **kwargs):
        resp = await _orig(self, **kwargs)
        try:
            um = getattr(resp, "usage_metadata", None)
            sink = _usage_sink.get()
            if um is not None and sink is not None:
                g = lambda a: int(getattr(um, a, 0) or 0)
                for k, a in (("input", "prompt_token_count"), ("output", "candidates_token_count"),
                             ("thinking", "thoughts_token_count"), ("total", "total_token_count")):
                    sink[k] = sink.get(k, 0) + g(a)
        except Exception:
            pass
        return resp

    AsyncModels.generate_content = _wrapped
    AsyncModels._bulk_tok_patched = True


async def process_one(t: Target, run_summary, retries: int, counter: dict, total: int,
                      report: "Report", annotator: "Optional[SourceAnnotator]" = None):
    """Summarize one doc, with in-run retries. Docs reach here already NOT_STARTED or
    FAILED (main() normalizes COMPLETED/IN_PROGRESS beforehand), so _run_document_summary
    will always process them. The DB row IS the state, so this is fully idempotent."""
    fid = uuid.UUID(t.file_id)
    sink: dict = {}                               # token usage accumulates here (across retries)
    tok_ctx = _usage_sink.set(sink)
    try:
        # Mirror the API endpoint: stamp a request id on the row before generating.
        try:
            await crud_document.update(fid, {"last_summary_request_id": uuid.uuid4()})
        except Exception:
            pass

        final = t.prior_status
        for attempt in range(1, retries + 2):    # 1 initial + `retries` extra
            t.attempts = attempt
            try:
                await asyncio.wait_for(run_summary(fid), timeout=PER_DOC_TIMEOUT)
            except asyncio.TimeoutError:
                await crud_document.update(fid, {"summary_status": "FAILED",
                                                 "summary_error": f"bulk runner timeout >{PER_DOC_TIMEOUT}s"})
            except Exception as e:                # _run_document_summary handles its own; be safe
                log.warning(f"[{t.file_id}] unexpected error (attempt {attempt}): {e}")
            doc = await crud_document.get_by_id(fid)
            final = doc.summary_status if doc else "UNKNOWN"
            t.error = doc.summary_error if doc else t.error
            if final == "COMPLETED":
                break                             # FAILED is not skipped, so a retry re-runs it
    finally:
        _usage_sink.reset(tok_ctx)

    t.tok_input, t.tok_output = sink.get("input", 0), sink.get("output", 0)
    t.tok_thinking, t.tok_total = sink.get("thinking", 0), sink.get("total", 0)
    t.final_status = final
    counter["done"] += 1
    toks = (f"  tokens[in={t.tok_input} out={t.tok_output} think={t.tok_thinking} total={t.tok_total}]"
            if t.tok_total else "")
    log.info(f"  ({counter['done']}/{total}) {t.file_id} {t.prior_status} -> {final}"
             + (f" [x{t.attempts}]" if t.attempts > 1 else "") + toks
             + (f"  ERROR: {t.error}" if final != "COMPLETED" else ""))
    await report.add(t)
    if annotator is not None:
        await annotator.maybe_write()


async def run_execute(to_process: list[Target], batch_size: int, retries: int, report: "Report",
                      annotator: "Optional[SourceAnnotator]" = None):
    # Import the REAL CBP logic only now (needs GCS + genai). Dry-run never reaches here.
    from src.api.v1.document_routes import _run_document_summary
    _install_token_logging()                      # capture per-doc token usage
    sem = asyncio.Semaphore(batch_size)
    counter = {"done": 0}
    total = len(to_process)

    async def _guarded(t: Target):
        async with sem:                           # keeps exactly `batch_size` in flight
            await process_one(t, _run_document_summary, retries, counter, total, report, annotator)

    # return_exceptions=True: one crashing worker never cancels the rest of the batch.
    await asyncio.gather(*[_guarded(t) for t in to_process], return_exceptions=True)


# ── plan / report ────────────────────────────────────────────────────────────
_FIELDS = ["sheet", "row", "resolution", "file_id", "filename", "state_id", "dept_id",
           "state_name", "dept_name", "prior_status", "final_status", "attempts",
           "tok_input", "tok_output", "tok_thinking", "tok_total", "error"]


def _row_of(t: Target) -> dict:
    return {"sheet": t.sheet or "", "row": t.excel_row, "resolution": t.resolution,
            "file_id": t.file_id or "", "filename": t.filename or "",
            "state_id": t.state_id or "", "dept_id": t.dept_id or "",
            "state_name": t.names[0], "dept_name": t.names[1],
            "prior_status": t.prior_status or "", "final_status": t.final_status or "",
            "attempts": t.attempts or "",
            "tok_input": t.tok_input or "", "tok_output": t.tok_output or "",
            "tok_thinking": t.tok_thinking or "", "tok_total": t.tok_total or "",
            "error": t.error or ""}


class Report:
    """Durable, incrementally-flushed CSV report. Rows are written as each doc finishes
    (flushed to disk), so a crash mid-run still leaves a usable partial report."""
    def __init__(self, path: str):
        self.path = path
        self._f = open(path, "w", newline="", encoding="utf-8")
        self._w = csv.DictWriter(self._f, fieldnames=_FIELDS)
        self._w.writeheader()
        self._f.flush()
        self._lock = asyncio.Lock()

    def add_sync(self, t: Target):
        self._w.writerow(_row_of(t))
        self._f.flush()

    async def add(self, t: Target):
        async with self._lock:
            self.add_sync(t)

    def close(self):
        self._f.close()


def write_report(path: str, targets: list[Target]):
    r = Report(path)
    for t in targets:
        r.add_sync(t)
    r.close()
    log.info(f"report written: {path}")


_STATUS_COLS = ["run_status", "run_docs_total", "run_docs_done", "run_docs_failed",
                "run_file_ids", "run_error", "run_updated_at"]


class SourceAnnotator:
    """Writes per-source-row status back into a CSV so the input doubles as a progress tracker.
    When the source is a .csv it overwrites the SAME file (atomically); for an .xlsx it writes a
    sibling '<name>.status.csv'. In SCOPE mode a row aggregates its scope's documents:
      run_status: COMPLETED | PARTIAL | FAILED | PENDING | UNRESOLVED | AMBIGUOUS | SKIPPED
    Refreshed incrementally during a run and once more at the end."""
    def __init__(self, path, headers, rows, targets, mode, cols, include_origin):
        self.path = path
        # drop any status/origin columns from a previous run so we don't duplicate headers
        self.headers = [h for h in headers if h not in _STATUS_COLS and h not in ("sheet", "row")]
        self.rows = rows
        self.mode = mode
        self.cols = cols
        self.include_origin = include_origin
        self._n = 0
        self._lock = asyncio.Lock()
        self.by_scope: dict[tuple, list[Target]] = {}
        self.by_rowkey: dict[tuple, list[Target]] = {}
        for t in targets:
            self.by_scope.setdefault((t.state_id, t.dept_id), []).append(t)
            self.by_rowkey.setdefault((t.sheet, t.excel_row), []).append(t)

    def _targets_for(self, r: dict) -> list[Target]:
        if self.mode == "scope":
            sid = _s(r.get(self.cols.get("state"))) if self.cols.get("state") else None
            did = _s(r.get(self.cols.get("dept"))) if self.cols.get("dept") else None
            return self.by_scope.get((sid, did), [])
        return self.by_rowkey.get((r.get(_SHEET_KEY, ""), r.get(_ROW_KEY)), [])

    def _agg(self, r: dict) -> dict:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        base = {c: "" for c in _STATUS_COLS}
        base["run_updated_at"] = now
        ts = self._targets_for(r)
        if not ts:
            base["run_status"] = "SKIPPED"      # row not selected (e.g. document_available != yes)
            return base
        matched = [t for t in ts if t.resolution == "matched"]
        if not matched:                          # unresolved / ambiguous scope
            base["run_status"] = ts[0].resolution.upper()
            base["run_error"] = "; ".join(sorted({t.error for t in ts if t.error}))
            return base
        done = [t for t in matched if (t.final_status or t.prior_status) == "COMPLETED"]
        failed = [t for t in matched if (t.final_status or "") == "FAILED"]
        base["run_status"] = ("COMPLETED" if len(done) == len(matched) else
                              "PARTIAL" if done else "FAILED" if failed else "PENDING")
        base["run_docs_total"] = len(matched)
        base["run_docs_done"] = len(done)
        base["run_docs_failed"] = len(failed)
        base["run_file_ids"] = ";".join(t.file_id for t in matched if t.file_id)
        base["run_error"] = "; ".join(sorted({t.error for t in matched if t.error}))
        return base

    def write(self):
        cols = (["sheet", "row"] if self.include_origin else []) + self.headers + _STATUS_COLS
        tmp = self.path + ".tmp"
        with open(tmp, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in self.rows:
                row = {h: (r.get(h) if r.get(h) is not None else "") for h in self.headers}
                if self.include_origin:
                    row["sheet"] = r.get(_SHEET_KEY, "")
                    row["row"] = r.get(_ROW_KEY, "")
                row.update(self._agg(r))
                w.writerow(row)
        os.replace(tmp, self.path)               # atomic, so a crash never leaves a half-written file

    async def maybe_write(self, every: int = 10):
        async with self._lock:
            self._n += 1
            if self._n % every == 0:
                self.write()


def _tally(items, attr):
    out: dict[str, int] = {}
    for it in items:
        out[getattr(it, attr) or "-"] = out.get(getattr(it, attr) or "-", 0) + 1
    return dict(sorted(out.items(), key=lambda kv: -kv[1]))


def _is_stale(doc: Document, minutes: int) -> bool:
    """True if an IN_PROGRESS doc hasn't been touched for `minutes` -> assume crashed."""
    ua = getattr(doc, "updated_at", None)
    if ua is None:
        return True
    if ua.tzinfo is None:
        ua = ua.replace(tzinfo=timezone.utc)
    return (datetime.now(timezone.utc) - ua) > timedelta(minutes=minutes)


async def main():
    ap = argparse.ArgumentParser(description="Bulk document-summary runner (dev workaround).")
    ap.add_argument("--excel", default=EXCEL_PATH, required=not EXCEL_PATH,
                    help="source file: .xlsx (all tabs) or .csv")
    ap.add_argument("--user-id", required=True, type=uuid.UUID,
                    help="Target user UUID -- every document lookup is restricted to "
                         "documents.uploader_id == this value. Mandatory.")
    ap.add_argument("--sheet", default=EXCEL_SHEET,
                    help="xlsx: restrict to one worksheet tab (default: read all tabs)")
    ap.add_argument("--csv", default=CSV_PATH, help="state/dept id->name reference (optional)")
    ap.add_argument("--out", default=OUT_PATH)
    ap.add_argument("--execute", action="store_true", help="actually generate summaries (default: dry run)")
    ap.add_argument("--limit", type=int, default=0, help="process at most N matched docs (0 = all)")
    ap.add_argument("--force", action="store_true", help="regenerate even if already COMPLETED")
    ap.add_argument("--reset-inprogress", action="store_true",
                    help="re-run ALL docs currently IN_PROGRESS (regardless of age)")
    ap.add_argument("--no-resume", action="store_true",
                    help="disable auto crash-recovery of stale IN_PROGRESS docs")
    ap.add_argument("--stale-minutes", type=int, default=RESUME_STALE_MINUTES,
                    help="IN_PROGRESS docs untouched this long are treated as crashed and re-run")
    ap.add_argument("--retries", type=int, default=RETRIES, help="extra in-run attempts on FAILED")
    ap.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    ap.add_argument("--filter-col", default=FILTER_COL)
    ap.add_argument("--filter-value", default="", help="override the kept value (default: 'yes')")
    ap.add_argument("--file-id-col", default=FILE_ID_COL)
    ap.add_argument("--filename-col", default=FILENAME_COL)
    ap.add_argument("--state-col", default=STATE_COL)
    ap.add_argument("--dept-col", default=DEPT_COL)
    ap.add_argument("--log-file", default=DEFAULT_LOG_FILE,
                    help="append-only log file (default: scripts/logs/bulk_summary_runner.log)")
    ap.add_argument("--status-out", default="",
                    help="where to write per-row status (default: the source .csv in place, "
                         "or '<source>.status.csv' for an .xlsx)")
    ap.add_argument("--no-annotate", action="store_true",
                    help="do not write per-row status back into a CSV")
    args = ap.parse_args()

    log_path = setup_logging(args.log_file)
    started = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    log.info("=" * 78)
    log.info(f"RUN {started} | mode={'EXECUTE' if args.execute else 'DRY-RUN'} | source={args.excel} "
             f"| user_id={args.user_id} | limit={args.limit or 'all'} | batch={args.batch_size} "
             f"| force={args.force} | resume_stale={not args.no_resume}@{args.stale_minutes}m "
             f"| retries={args.retries}")
    log.info(f"logging to: {log_path}")

    sessionmanager.init(settings.DATABASE_URL)
    log.info(f"DB: {str(settings.DATABASE_URL).split('@')[-1]}  |  storage: {settings.DOCUMENT_STORAGE_TYPE}")

    # ── load + detect columns ──
    headers, rows = load_rows(args.excel, args.sheet)
    log.info(f"source: {len(rows)} data rows, headers: {headers}")

    filter_col = detect_col(headers, args.filter_col, "filter")
    file_id_col = detect_col(headers, args.file_id_col, "file_id", require_uuid_sample=rows)
    filename_col = detect_col(headers, args.filename_col, "filename")
    state_col = detect_col(headers, args.state_col, "state")
    dept_col = detect_col(headers, args.dept_col, "dept")
    statename_col = detect_col(headers, "", "statename")
    deptname_col = detect_col(headers, "", "deptname")
    ndocs_col = detect_col(headers, "", "ndocs")

    # per-document mode if the sheet carries a document id; else scope mode (fetch from DB).
    mode = "file_id" if file_id_col else ("filename" if filename_col else "scope")
    log.info("detected columns -> " + ", ".join(
        f"{k}={v!r}" for k, v in [("filter", filter_col), ("file_id", file_id_col),
                                  ("filename", filename_col), ("state", state_col), ("dept", dept_col),
                                  ("state_name", statename_col), ("dept_name", deptname_col),
                                  ("n_docs", ndocs_col)]))
    log.info(f"resolution mode: {mode.upper()}"
             + ("  (each 'yes' row -> that state/dept's documents are fetched from the DB)"
                if mode == "scope" else ""))
    if mode == "scope" and not (state_col and dept_col):
        raise SystemExit("[config] no file_id/filename column AND no state+dept id columns -- "
                         "cannot resolve documents. Pass --state-col/--dept-col or --file-id-col.")

    # ── filter to the "yes" rows (no filter column -> every row is processed) ──
    if filter_col:
        keep_vals = {args.filter_value.lower()} if args.filter_value else FILTER_TRUE_VALUES
        filtered = [r for r in rows if _s(r.get(filter_col)) and str(r.get(filter_col)).strip().lower() in keep_vals]
        log.info(f"filter {filter_col!r} in {keep_vals}: {len(filtered)}/{len(rows)} rows kept")
    else:
        filtered = rows
        log.info("no filter column detected -> processing all rows")
    if not filtered:
        raise SystemExit("nothing matched the filter -- check --filter-col / --filter-value.")

    # ── guard: id columns must not be scientific-notation/float (Excel corrupts long ids) ──
    for col, label in [(state_col, "state"), (dept_col, "dept")]:
        if not col:
            continue
        vals = [_s(r.get(col)) for r in filtered if _s(r.get(col))]
        bad = [v for v in vals if re.search(r"[eE]\+?\d|\.\d", v)]
        if bad:
            log.warning(f"!! {label} id column {col!r}: {len(bad)}/{len(vals)} values look NUMERIC "
                        f"(e.g. {bad[0]!r}) -- the real long ids are likely CORRUPTED by Excel, so "
                        f"those DB matches will fail. Format the column as TEXT in the xlsx, or "
                        f"supply a CSV with ids kept as strings.")

    # ── resolve to DB documents ──
    cols = {"file_id": file_id_col, "filename": filename_col, "state": state_col, "dept": dept_col,
            "statename": statename_col, "deptname": deptname_col, "ndocs": ndocs_col}
    if mode == "scope":
        n_scopes = len({(_s(r.get(state_col)), _s(r.get(dept_col))) for r in filtered})
        log.info(f"{n_scopes} unique (state, dept) scope(s) from {len(filtered)} kept rows")
        targets = await resolve_scope_targets(filtered, cols, args.user_id)
    else:
        targets = await resolve_targets(filtered, cols, args.user_id)

    # ── enrich names from the optional CSV (only where the sheet didn't already provide them) ──
    if args.csv:
        chdrs, crows = load_csv_rows(args.csv)
        csid = detect_col(chdrs, CSV_STATE_ID_COL, "state")
        cdid = detect_col(chdrs, CSV_DEPT_ID_COL, "dept")
        csn = detect_col(chdrs, CSV_STATE_NAME_COL, "statename")
        cdn = detect_col(chdrs, CSV_DEPT_NAME_COL, "deptname")
        name_map: dict[tuple, tuple] = {}
        for cr in crows:
            key = (_s(cr.get(csid)) if csid else None, _s(cr.get(cdid)) if cdid else None)
            name_map[key] = (_s(cr.get(csn)) if csn else "", _s(cr.get(cdn)) if cdn else "")
        for t in targets:
            if t.names == ("", ""):
                t.names = name_map.get((t.state_id, t.dept_id)) or name_map.get((t.state_id, None)) or ("", "")

    # ── per-row status writer (source csv in place, or a sibling .status.csv for xlsx) ──
    annotator = None
    if not args.no_annotate:
        is_csv = args.excel.lower().endswith(".csv")
        status_out = args.status_out or (args.excel if is_csv
                                         else str(Path(args.excel).with_suffix("")) + ".status.csv")
        annotator = SourceAnnotator(status_out, headers, rows, targets, mode, cols,
                                    include_origin=not is_csv)
        log.info(f"per-row status -> {status_out}")

    # ── plan report ──
    matched = [t for t in targets if t.resolution == "matched"]
    log.info(f"resolution: {_tally(targets, 'resolution')}")
    log.info(f"matched by prior status: {_tally(matched, 'prior_status')}")

    # Which matched docs will we (re)process? The DB row's summary_status is the state:
    #   NOT_STARTED / FAILED  -> always (re)process   (crash-restart & retry are automatic)
    #   COMPLETED             -> skip, unless --force
    #   IN_PROGRESS           -> skip (assume a live run owns it), UNLESS it's stale
    #                           (crashed) or --reset-inprogress
    def selectable(t: Target) -> bool:
        p = t.prior_status
        if p in ("NOT_STARTED", "FAILED"):
            return True
        if p == "COMPLETED":
            return args.force
        if p == "IN_PROGRESS":
            return args.reset_inprogress or (not args.no_resume and _is_stale(t.doc, args.stale_minutes))
        return False

    to_process = [t for t in matched if selectable(t)]
    if args.limit:
        to_process = to_process[:args.limit]
    recovered = [t for t in to_process if t.prior_status == "IN_PROGRESS"]
    log.info(f"TO PROCESS: {len(to_process)} docs  (force={args.force}, "
             f"reset_inprogress={args.reset_inprogress}, resume_stale={not args.no_resume}"
             f"@{args.stale_minutes}m -> {len(recovered)} recovered, limit={args.limit or 'none'})")

    out = args.out or str(Path(args.excel).with_name(
        f"bulk_summary_{'run' if args.execute else 'plan'}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"))

    if not args.execute:
        write_report(out, targets)
        if annotator is not None:
            annotator.write()
            log.info(f"per-row status written: {annotator.path}")
        log.info("DRY RUN complete -- no summaries generated. Re-run with --execute to process.")
        await sessionmanager.close()
        return

    # ── normalize DB state so _run_document_summary won't skip a doc we intend to run ──
    # (it skips COMPLETED/IN_PROGRESS). Do this up front and in-memory-mark prior_status.
    for t in to_process:
        fid = uuid.UUID(t.file_id)
        if t.prior_status == "COMPLETED":         # --force
            await crud_document.update(fid, {"summary_status": "NOT_STARTED",
                                             "summary_text": None, "summary_error": None})
        elif t.prior_status == "IN_PROGRESS":     # recovered stale / --reset-inprogress
            await crud_document.update(fid, {"summary_status": "NOT_STARTED", "summary_error": None})

    # ── execute (incremental, durable report; write skipped rows first) ──
    report = Report(out)
    proc_ids = {id(t) for t in to_process}
    for t in targets:                             # everything we are NOT processing, recorded up front
        if id(t) not in proc_ids:
            t.final_status = t.final_status or t.prior_status
            report.add_sync(t)
    if annotator is not None:
        annotator.write()                         # initial snapshot (skipped/unresolved rows visible)
    log.info(f"EXECUTING {len(to_process)} summaries, {args.batch_size} at a time "
             f"(retries={args.retries}, per-doc timeout={PER_DOC_TIMEOUT}s) ...")
    try:
        await run_execute(to_process, args.batch_size, args.retries, report, annotator)
    finally:
        report.close()
        if annotator is not None:
            annotator.write()                     # final refresh, even if interrupted
    log.info(f"final status of processed docs: {_tally(to_process, 'final_status')}")
    tin = sum(t.tok_input for t in to_process)
    tout = sum(t.tok_output for t in to_process)
    tthink = sum(t.tok_thinking for t in to_process)
    ttot = sum(t.tok_total for t in to_process)
    log.info(f"TOTAL tokens over {len(to_process)} docs: input={tin} output={tout} "
             f"thinking={tthink} total={ttot}")
    log.info(f"report: {out}")
    if annotator is not None:
        log.info(f"per-row status: {annotator.path}")
    await sessionmanager.close()


if __name__ == "__main__":
    asyncio.run(main())