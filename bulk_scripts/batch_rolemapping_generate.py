"""
Bulk single-designation role-mapping runner (backend workaround, run in DEV).

Companion to scripts/bulk_summary_runner.py -- SAME input file (.xlsx multi-tab / .csv),
SAME harness (dedup, batch-of-10, idempotent resume, append logging, token capture,
per-row status write-back, dry-run). Instead of summaries it generates a v3 role mapping
for ONE designation per row.

It reuses the v3 role-mapping logic (src/services/v3/role_mapping_service.py) but runs only
the passes that matter for a single, already-known designation:
    PASS 1 (designation extraction)  -> SKIPPED (the designation is given in the file)
    PASS 2 (FRAC generation)         -> RUN     (1 LLM call per designation)
    PASS 3 (domain-from-WAO)         -> SKIPPED (not required)
    PASS 4 (KCM reconciliation)      -> RUN     (deterministic, no LLM)
then it saves one `role_mappings` row (status=COMPLETED), exactly like the v3 flow -- WITHOUT
the API's extra bookkeeping (placeholder rows, multi-row split, iGOT matching). iGOT matching
is available as an opt-in (--igot-match).

Prerequisites:
  * Document summaries must already exist for each scope (run bulk_summary_runner.py first) --
    PASS 2 reads them. A scope with no COMPLETED summaries is reported `unresolved`.
  * .env points at the dev DB + GCS bucket + Vertex creds (all required -- the app's Settings
    object fails fast at import if any are missing).

Mandatory CLI args: --excel and --user-id. There is no dry-run/execute default fallback for
either -- both must always be passed.

Source file columns (ALL mandatory -- missing any of them is a fatal error):
    state_center_id, department_id, org_type, state_center_name, department_name, designation
There is no filter/yes-no column in this file format -- every data row is processed.
org_type must parse to 'state' or 'ministry' (see _org_type_of); an unparseable value marks
just that row `unresolved` (with an error) rather than aborting the whole run.

USAGE:
  # dry run (no LLM): resolve scopes, check summaries exist, show plan, write status into the CSV
  .venv/bin/python scripts/role_mapping_runner.py --excel /path/source.csv --user-id <UUID>

  # smoke test 5, then full run
  .venv/bin/python scripts/role_mapping_runner.py --excel /path/source.csv --user-id <UUID> --execute --limit 5
  .venv/bin/python scripts/role_mapping_runner.py --excel /path/source.csv --user-id <UUID> --execute
"""
from __future__ import annotations

import argparse
import asyncio
import contextvars
import csv
import logging
import os
import pkgutil
import re
import sys
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

# ── path bootstrap ────────────────────────────────────────────────────────────
_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(_ROOT))

from src.core.configs import settings  # noqa: E402  (loads .env)

if settings.GOOGLE_APPLICATION_CREDENTIALS:
    os.environ.setdefault("GOOGLE_APPLICATION_CREDENTIALS", settings.GOOGLE_APPLICATION_CREDENTIALS)

# Register ALL SQLAlchemy mappers (the server imports every model at startup).
import src.models  # noqa: E402
for _m in pkgutil.iter_modules(src.models.__path__):
    __import__(f"src.models.{_m.name}")

from sqlalchemy import select, and_, delete, func  # noqa: E402
from src.core.database import sessionmanager  # noqa: E402
from src.models.document import Document  # noqa: E402
from src.models.role_mapping import RoleMapping, ProcessingStatus  # noqa: E402
from src.schemas.role_mapping import OrgType  # noqa: E402
from src.crud.role_mapping import crud_role_mapping  # noqa: E402

# Import the app logger up front: it runs logging.config.fileConfig() which disables existing
# loggers + replaces handlers. Triggering it now (cached) means our setup_logging() sticks.
import src.core.logger  # noqa: E402,F401

log = logging.getLogger("rolemap_runner")

_usage_sink: contextvars.ContextVar = contextvars.ContextVar("rm_usage", default=None)
# default log file -- one fresh, timestamped file per run (matches the other bulk_scripts)
RUN_TIMESTAMP = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
DEFAULT_LOG_FILE = str(Path(__file__).resolve().parent / "logs" / f"role_mapping_runner_{RUN_TIMESTAMP}.log")


def setup_logging(path: str) -> str:
    """Console + file (one fresh file per run). Installed after the app's fileConfig."""
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    root = logging.getLogger()
    for h in list(root.handlers):
        root.removeHandler(h)
    con = logging.StreamHandler()
    con.setFormatter(logging.Formatter("%(asctime)s %(levelname)s %(message)s"))
    fh = logging.FileHandler(path, mode="w", encoding="utf-8")
    fh.setFormatter(logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s"))
    root.addHandler(con)
    root.addHandler(fh)
    root.setLevel(logging.INFO)
    for name in ("rolemap_runner", "ai_cbp_service"):
        lg = logging.getLogger(name)
        lg.disabled = False
        lg.handlers = []
        lg.propagate = True
        lg.setLevel(logging.INFO)
    logging.getLogger("google").setLevel(logging.WARNING)
    return path


# ══════════════════════════════════════════════════════════════════════════════
# CONFIG (all overridable by CLI flags)
# ══════════════════════════════════════════════════════════════════════════════
EXCEL_PATH = ""
EXCEL_SHEET = ""
OUT_PATH = ""
BATCH_SIZE = 10
PER_DOC_TIMEOUT = 1200
RETRIES = 1

_CAND = {
    "state":       {"statecenterid", "statecenter", "stateid", "state", "orgid", "frameworkid"},
    "dept":        {"departmentmdoid", "mdoid", "deptmdoid", "departmentid", "deptid",
                    "department", "dept"},
    "statename":   {"statecentername", "statename"},
    "deptname":    {"departmentmdoname", "mdoname", "departmentname", "deptname"},
    "designation": {"designationname", "designation", "role", "post", "jobtitle"},
    "orgtype":     {"orgtype", "organizationtype", "organisationtype", "type", "level", "category"},
}


def _norm(s) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower()) if s is not None else ""


def _s(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _org_type_of(val) -> Optional[OrgType]:
    """Parses a row's org_type value. Returns None (not a fallback) if it doesn't match
    a known state/ministry synonym -- org_type is a mandatory column with no CLI default,
    so an unparseable value must be surfaced as an error, not silently defaulted."""
    n = _norm(val)
    if n in ("state", "states"):
        return OrgType.state
    if n in ("ministry", "ministries", "centre", "center", "central", "union"):
        return OrgType.ministry
    return None


# ── source loading (.xlsx multi-tab / .csv) ──────────────────────────────────
_SHEET_KEY = "__sheet__"
_ROW_KEY = "__row__"


def load_csv_rows(path: str) -> tuple[list[str], list[dict]]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return (reader.fieldnames or []), list(reader)


def load_rows(path: str, sheet: str = "") -> tuple[list[str], list[dict]]:
    """(headers, rows) from .csv or .xlsx (all tabs unless `sheet`). Rows carry origin keys."""
    if path.lower().endswith(".csv"):
        hdrs, rows = load_csv_rows(path)
        for i, r in enumerate(rows, start=2):
            r[_SHEET_KEY], r[_ROW_KEY] = "(csv)", i
        return hdrs, rows
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = [wb[sheet]] if sheet else list(wb.worksheets)
    all_headers, seen, rows, per_sheet = [], set(), [], []
    for ws in sheets:
        it = ws.iter_rows(values_only=True)
        try:
            first = next(it)
        except StopIteration:
            continue
        headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(first)]
        for h in headers:
            if h not in seen:
                seen.add(h)
                all_headers.append(h)
        cnt = 0
        for rnum, raw in enumerate(it, start=2):
            if raw is None or all(c is None for c in raw):
                continue
            row = {headers[i]: raw[i] if i < len(raw) else None for i in range(len(headers))}
            row[_SHEET_KEY], row[_ROW_KEY] = ws.title, rnum
            rows.append(row)
            cnt += 1
        per_sheet.append((ws.title, cnt))
    wb.close()
    if len(per_sheet) > 1:
        log.info(f"read {len(per_sheet)} tabs: " + ", ".join(f"{t}={c}" for t, c in per_sheet))
    return all_headers, rows


def _origin(r: dict, fallback_idx: int) -> tuple[str, int]:
    return r.get(_SHEET_KEY, ""), r.get(_ROW_KEY, fallback_idx)


def detect_col(headers: list[str], override: str, kind: str) -> Optional[str]:
    if override:
        m = next((h for h in headers if h == override or _norm(h) == _norm(override)), None)
        if not m:
            raise SystemExit(f"[config] column '{override}' (for {kind}) not found. Headers: {headers}")
        return m
    cands = _CAND.get(kind, set())
    return next((h for h in headers if _norm(h) in cands), None)


# ── DB helpers (scope summaries, existing mappings) ──────────────────────────
async def _scope_summary_count(user_id: uuid.UUID, state_id: str, dept_id: Optional[str]) -> int:
    """How many COMPLETED document summaries exist in this scope for this user (cheap
    availability check)."""
    conds = [Document.state_center_id == state_id, Document.summary_status == "COMPLETED",
             Document.uploader_id == user_id]
    conds.append(Document.department_id == dept_id if dept_id else Document.department_id.is_(None))
    async with sessionmanager.session() as db:
        return int((await db.execute(select(func.count(Document.file_id)).where(and_(*conds)))).scalar() or 0)


async def _scope_summary_text(user_id: uuid.UUID, state_id: str, dept_id: Optional[str]) -> str:
    """Full document summaries for a scope, formatted like the v3 service's get_documents_summary.
    Restricted to documents uploaded by user_id."""
    conds = [Document.state_center_id == state_id, Document.summary_status == "COMPLETED",
             Document.uploader_id == user_id]
    conds.append(Document.department_id == dept_id if dept_id else Document.department_id.is_(None))
    async with sessionmanager.session() as db:
        docs = (await db.execute(select(Document).where(and_(*conds)))).scalars().all()
    parts = []
    for idx, doc in enumerate(docs, start=1):
        summary = (doc.summary_text or "").strip()
        parts.append(f"<document_summary_{idx}>\n Document Type: {doc.document_type} \n "
                     f"Summary: {summary}\n</document_summary_{idx}>")
    return "\n\n".join(parts)


async def _existing_completed(user_id: uuid.UUID, state_id: str, dept_id: Optional[str]) -> set:
    """Set of designation_name (lowercased) that already have a COMPLETED role mapping for this
    (user, state, dept) -- used for idempotent skip/resume."""
    conds = [RoleMapping.user_id == user_id, RoleMapping.state_center_id == state_id,
             RoleMapping.status == ProcessingStatus.COMPLETED]
    conds.append(RoleMapping.department_id == dept_id if dept_id else RoleMapping.department_id.is_(None))
    async with sessionmanager.session() as db:
        rows = (await db.execute(select(RoleMapping.designation_name).where(and_(*conds)))).all()
    return {(_s(r[0]) or "").lower() for r in rows}


async def _delete_existing(user_id: uuid.UUID, state_id: str, dept_id: Optional[str], designation: str):
    conds = [RoleMapping.user_id == user_id, RoleMapping.state_center_id == state_id,
             func.lower(RoleMapping.designation_name) == (designation or "").lower()]
    conds.append(RoleMapping.department_id == dept_id if dept_id else RoleMapping.department_id.is_(None))
    async with sessionmanager.session() as db:
        await db.execute(delete(RoleMapping).where(and_(*conds)))
        await db.commit()


# scope summary cache (fetched once per scope, reused across its designations)
_summary_cache: dict[tuple, str] = {}
_summary_lock = asyncio.Lock()


async def _get_scope_summary(user_id: uuid.UUID, state_id: str, dept_id: Optional[str]) -> str:
    key = (state_id, dept_id)
    if key in _summary_cache:
        return _summary_cache[key]
    async with _summary_lock:
        if key not in _summary_cache:
            _summary_cache[key] = await _scope_summary_text(user_id, state_id, dept_id)
        return _summary_cache[key]


# ── target model + resolution ────────────────────────────────────────────────
@dataclass
class Target:
    excel_row: int
    sheet: str = ""
    state_id: Optional[str] = None
    dept_id: Optional[str] = None
    designation: Optional[str] = None
    org_type: OrgType = OrgType.state
    resolution: str = "pending"        # matched | unresolved | duplicate
    prior_status: Optional[str] = None  # COMPLETED (already has a mapping) | NOT_STARTED
    final_status: Optional[str] = None
    error: Optional[str] = None
    attempts: int = 0
    comp_count: int = 0
    role_mapping_id: Optional[str] = None
    tok_input: int = 0
    tok_output: int = 0
    tok_thinking: int = 0
    tok_total: int = 0
    names: tuple = field(default=("", ""))


async def resolve_targets(rows: list[dict], cols: dict, user_id: uuid.UUID) -> list[Target]:
    """One Target per unique (state, dept, designation). Resolution needs, per scope: whether
    COMPLETED summaries exist (else `unresolved`), and which designations already have a mapping."""
    seen: set = set()
    targets: list[Target] = []
    for idx, r in enumerate(rows, start=2):
        sid = _s(r.get(cols["state"])) if cols.get("state") else None
        did = _s(r.get(cols["dept"])) if cols.get("dept") else None
        desig = _s(r.get(cols["designation"])) if cols.get("designation") else None
        key = (sid, did, (desig or "").lower())
        sheet, rownum = _origin(r, idx)
        if desig is None:
            targets.append(Target(excel_row=rownum, sheet=sheet, state_id=sid, dept_id=did,
                                  resolution="unresolved", error="missing designation"))
            continue
        if key in seen:
            targets.append(Target(excel_row=rownum, sheet=sheet, state_id=sid, dept_id=did,
                                  designation=desig, resolution="duplicate"))
            continue
        seen.add(key)
        sn = _s(r.get(cols["statename"])) if cols.get("statename") else None
        dn = _s(r.get(cols["deptname"])) if cols.get("deptname") else None
        ot = _org_type_of(r.get(cols["orgtype"]))
        if ot is None:
            targets.append(Target(excel_row=rownum, sheet=sheet, state_id=sid, dept_id=did,
                                  designation=desig, resolution="unresolved",
                                  error=f"invalid/missing org_type value: {r.get(cols['orgtype'])!r} "
                                        f"(must be 'state' or 'ministry')"))
            continue
        targets.append(Target(excel_row=rownum, sheet=sheet, state_id=sid, dept_id=did,
                              designation=desig, org_type=ot, names=(sn or "", dn or "")))

    # per-scope: summary availability + existing completed designations
    scopes = {(t.state_id, t.dept_id) for t in targets if t.resolution == "pending"}
    summ_count: dict[tuple, int] = {}
    existing: dict[tuple, set] = {}
    for sc in scopes:
        summ_count[sc] = await _scope_summary_count(user_id, sc[0], sc[1])
        existing[sc] = await _existing_completed(user_id, sc[0], sc[1])

    for t in targets:
        if t.resolution != "pending":
            continue
        sc = (t.state_id, t.dept_id)
        if summ_count.get(sc, 0) == 0:
            t.resolution = "unresolved"
            t.error = "no COMPLETED document summaries in scope (run summary generation first)"
            continue
        t.resolution = "matched"
        t.prior_status = "COMPLETED" if (t.designation or "").lower() in existing.get(sc, set()) else "NOT_STARTED"
    return targets


# ── token capture ────────────────────────────────────────────────────────────
def _install_token_logging() -> None:
    try:
        from google.genai.models import AsyncModels
    except Exception as exc:
        log.warning(f"token logging unavailable: {exc}")
        return
    if getattr(AsyncModels, "_rm_tok_patched", False):
        return
    _orig = AsyncModels.generate_content

    async def _wrapped(self, **kwargs):
        resp = await _orig(self, **kwargs)
        try:
            um = getattr(resp, "usage_metadata", None)
            sink = _usage_sink.get()
            if um is not None and sink is not None:
                g = lambda a: int(getattr(um, a, 0) or 0)
                for k, a in (("input", "prompt_token_count"), ("output", "candidates_token_count"),
                             ("thinking", "thoughts_token_count"), ("total", "total_token_count")):
                    sink[k] = sink.get(k, 0) + g(a)
        except Exception:
            pass
        return resp

    AsyncModels.generate_content = _wrapped
    AsyncModels._rm_tok_patched = True


# ── execution ────────────────────────────────────────────────────────────────
async def process_one(t: Target, svc, matcher, user_id: uuid.UUID, instruction: Optional[str],
                      force: bool, retries: int, counter: dict, total: int,
                      report: "Report", annotator: "Optional[SourceAnnotator]"):
    """PASS 2 (FRAC) + PASS 4 (KCM reconcile) for one designation, then save one role_mappings row."""
    sink: dict = {}
    tok_ctx = _usage_sink.set(sink)
    final, mapping = "FAILED", None
    try:
        summary = await _get_scope_summary(user_id, t.state_id, t.dept_id) or "N/A"
        org_data = {
            "org_type": t.org_type.value,
            "state_center_id": t.state_id,
            "department_id": t.dept_id,
            "organization_name": t.names[0] or t.state_id,
            "department_name": t.names[1] or "N/A",
            "docs_summary": summary if summary else "N/A",
            "instruction": instruction or "N/A",
        }
        batch = [{"designation": t.designation, "sort_order": 1, "wing_division_section": "N/A"}]
        for attempt in range(1, retries + 2):
            t.attempts = attempt
            try:
                res = await asyncio.wait_for(
                    svc._generate_frac_for_batch(batch, org_data, batch_number=1),
                    timeout=PER_DOC_TIMEOUT)
                if res:
                    mapping = svc.reconcile_role_mappings_with_kcm(res)[0]   # PASS 4 (sync)
                    final = "COMPLETED"
                    break
                t.error = "empty FRAC response from model"
            except asyncio.TimeoutError:
                t.error = f"timeout >{PER_DOC_TIMEOUT}s"
            except Exception as e:
                t.error = str(e)
                log.warning(f"[{t.state_id}/{t.designation}] error (attempt {attempt}): {e}")
    finally:
        _usage_sink.reset(tok_ctx)

    if final == "COMPLETED" and mapping is not None:
        try:
            if force:
                await _delete_existing(user_id, t.state_id, t.dept_id, t.designation)
            row = RoleMapping(
                user_id=user_id, org_type=t.org_type.value,
                state_center_id=t.state_id, department_id=t.dept_id or None,
                state_center_name=t.names[0] or None, department_name=t.names[1] or None,
                instruction=instruction, status=ProcessingStatus.COMPLETED,
                designation_name=mapping.get("designation_name") or t.designation,
                wing_division_section=mapping.get("wing_division_section"),
                role_responsibilities=mapping.get("role_responsibilities") or [],
                activities=mapping.get("activities") or [],
                competencies=mapping.get("competencies") or [],
                sort_order=mapping.get("sort_order"),
            )
            created = await crud_role_mapping.create([row])
            t.role_mapping_id = str(created[0].id)
            t.comp_count = len(mapping.get("competencies") or [])
            if matcher is not None:
                try:
                    async with sessionmanager.session() as db:
                        res = await matcher.match(db, [t.designation])
                    m = next((x for x in (res or []) if (x.get("input_designation") or "").lower()
                              == (t.designation or "").lower()), None)
                    if m and m.get("id"):
                        await crud_role_mapping.update(created[0].id, {
                            "igot_designation_name": m.get("designation"),
                            "igot_designation_id": m.get("id")})
                except Exception as e:
                    log.warning(f"iGOT match failed for '{t.designation}' (non-critical): {e}")
        except Exception as e:
            final, t.error = "FAILED", f"save failed: {e}"
            log.exception(f"save failed for {t.state_id}/{t.designation}")

    t.tok_input, t.tok_output = sink.get("input", 0), sink.get("output", 0)
    t.tok_thinking, t.tok_total = sink.get("thinking", 0), sink.get("total", 0)
    t.final_status = final
    counter["done"] += 1
    toks = (f"  tokens[in={t.tok_input} out={t.tok_output} think={t.tok_thinking} total={t.tok_total}]"
            if t.tok_total else "")
    log.info(f"  ({counter['done']}/{total}) {t.state_id}/{t.dept_id or '-'} :: {t.designation} "
             f"{t.prior_status} -> {final}"
             + (f" [x{t.attempts}]" if t.attempts > 1 else "")
             + (f" comps={t.comp_count}" if final == "COMPLETED" else "") + toks
             + (f"  ERROR: {t.error}" if final != "COMPLETED" else ""))
    await report.add(t)
    if annotator is not None:
        await annotator.maybe_write()


async def run_execute(to_process: list[Target], user_id: uuid.UUID, instruction: Optional[str],
                      force: bool, igot_match: bool, batch_size: int, retries: int,
                      report: "Report", annotator: "Optional[SourceAnnotator]"):
    from src.services.v3.role_mapping_service import role_mapping_service as svc
    matcher = None
    if igot_match:
        from src.services.designation_matcher_service import designation_matcher_service as matcher
    _install_token_logging()
    sem = asyncio.Semaphore(batch_size)
    counter = {"done": 0}
    total = len(to_process)

    async def _guarded(t: Target):
        async with sem:
            await process_one(t, svc, matcher, user_id, instruction, force, retries,
                              counter, total, report, annotator)

    await asyncio.gather(*[_guarded(t) for t in to_process], return_exceptions=True)


# ── report + per-row status ──────────────────────────────────────────────────
_FIELDS = ["sheet", "row", "resolution", "state_id", "dept_id", "designation", "org_type",
           "state_name", "dept_name", "prior_status", "final_status", "attempts",
           "competencies", "role_mapping_id", "tok_input", "tok_output", "tok_thinking",
           "tok_total", "error"]


def _row_of(t: Target) -> dict:
    return {"sheet": t.sheet or "", "row": t.excel_row, "resolution": t.resolution,
            "state_id": t.state_id or "", "dept_id": t.dept_id or "",
            "designation": t.designation or "", "org_type": t.org_type.value if t.org_type else "",
            "state_name": t.names[0], "dept_name": t.names[1],
            "prior_status": t.prior_status or "", "final_status": t.final_status or "",
            "attempts": t.attempts or "", "competencies": t.comp_count or "",
            "role_mapping_id": t.role_mapping_id or "",
            "tok_input": t.tok_input or "", "tok_output": t.tok_output or "",
            "tok_thinking": t.tok_thinking or "", "tok_total": t.tok_total or "",
            "error": t.error or ""}


class Report:
    def __init__(self, path: str):
        self.path = path
        self._f = open(path, "w", newline="", encoding="utf-8")
        self._w = csv.DictWriter(self._f, fieldnames=_FIELDS)
        self._w.writeheader()
        self._f.flush()
        self._lock = asyncio.Lock()

    def add_sync(self, t: Target):
        self._w.writerow(_row_of(t))
        self._f.flush()

    async def add(self, t: Target):
        async with self._lock:
            self.add_sync(t)

    def close(self):
        self._f.close()


def write_report(path: str, targets: list[Target]):
    r = Report(path)
    for t in targets:
        r.add_sync(t)
    r.close()
    log.info(f"report written: {path}")


_STATUS_COLS = ["run_status", "run_competencies", "run_role_mapping_id", "run_tokens",
                "run_error", "run_updated_at"]


class SourceAnnotator:
    """Per-row status written back into the source CSV (in place for a .csv; sibling for .xlsx),
    so the input doubles as a progress tracker. One Target per (state, dept, designation), so
    each source row maps to exactly one Target (by sheet+row)."""
    def __init__(self, path, headers, rows, targets, cols, include_origin):
        self.path = path
        self.headers = [h for h in headers if h not in _STATUS_COLS and h not in ("sheet", "row")]
        self.rows = rows
        self.cols = cols
        self.include_origin = include_origin
        self._n = 0
        self._lock = asyncio.Lock()
        self.by_rowkey: dict[tuple, Target] = {}
        for t in targets:
            self.by_rowkey[(t.sheet, t.excel_row)] = t

    def _agg(self, r: dict) -> dict:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        base = {c: "" for c in _STATUS_COLS}
        base["run_updated_at"] = now
        t = self.by_rowkey.get((r.get(_SHEET_KEY, ""), r.get(_ROW_KEY)))
        if t is None:
            base["run_status"] = "SKIPPED"
            return base
        if t.resolution != "matched":
            base["run_status"] = t.resolution.upper()
            base["run_error"] = t.error or ""
            return base
        base["run_status"] = t.final_status or "PENDING"
        base["run_competencies"] = t.comp_count or ""
        base["run_role_mapping_id"] = t.role_mapping_id or ""
        base["run_tokens"] = t.tok_total or ""
        base["run_error"] = t.error or ""
        return base

    def write(self):
        cols = (["sheet", "row"] if self.include_origin else []) + self.headers + _STATUS_COLS
        tmp = self.path + ".tmp"
        with open(tmp, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            w.writeheader()
            for r in self.rows:
                row = {h: (r.get(h) if r.get(h) is not None else "") for h in self.headers}
                if self.include_origin:
                    row["sheet"], row["row"] = r.get(_SHEET_KEY, ""), r.get(_ROW_KEY, "")
                row.update(self._agg(r))
                w.writerow(row)
        os.replace(tmp, self.path)

    async def maybe_write(self, every: int = 10):
        async with self._lock:
            self._n += 1
            if self._n % every == 0:
                self.write()


def _tally(items, attr):
    out: dict[str, int] = {}
    for it in items:
        out[getattr(it, attr) or "-"] = out.get(getattr(it, attr) or "-", 0) + 1
    return dict(sorted(out.items(), key=lambda kv: -kv[1]))


async def main():
    ap = argparse.ArgumentParser(description="Bulk single-designation role-mapping runner (dev).")
    ap.add_argument("--excel", default=EXCEL_PATH, required=not EXCEL_PATH,
                    help="source file: .xlsx (all tabs) or .csv")
    ap.add_argument("--sheet", default=EXCEL_SHEET, help="xlsx: restrict to one tab (default: all)")
    ap.add_argument("--user-id", required=True, type=uuid.UUID,
                    help="UUID to attribute generated role mappings to. Mandatory.")
    ap.add_argument("--instruction", default="", help="optional extra instruction for generation")
    ap.add_argument("--igot-match", default="on", action="store_true", help="also match designations to the iGOT master")
    ap.add_argument("--out", default=OUT_PATH)
    ap.add_argument("--execute", action="store_true", help="actually generate (default: dry run)")
    ap.add_argument("--limit", type=int, default=0, help="process at most N (0 = all)")
    ap.add_argument("--force", action="store_true", help="regenerate even if a mapping already exists")
    ap.add_argument("--batch-size", type=int, default=BATCH_SIZE)
    ap.add_argument("--retries", type=int, default=RETRIES)
    ap.add_argument("--state-col", default="")
    ap.add_argument("--dept-col", default="")
    ap.add_argument("--designation-col", default="")
    ap.add_argument("--log-file", default=DEFAULT_LOG_FILE)
    ap.add_argument("--status-out", default="")
    ap.add_argument("--no-annotate", action="store_true")
    args = ap.parse_args()

    log_path = setup_logging(args.log_file)
    started = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    log.info("=" * 78)
    log.info(f"RUN {started} | mode={'EXECUTE' if args.execute else 'DRY-RUN'} | source={args.excel} "
             f"| user_id={args.user_id} | igot_match={args.igot_match} | limit={args.limit or 'all'} "
             f"| batch={args.batch_size}")
    log.info(f"logging to: {log_path}")

    user_uuid: uuid.UUID = args.user_id

    sessionmanager.init(settings.DATABASE_URL)
    log.info(f"DB: {str(settings.DATABASE_URL).split('@')[-1]}")

    headers, rows = load_rows(args.excel, args.sheet)
    log.info(f"source: {len(rows)} data rows, headers: {headers}")

    state_col = detect_col(headers, args.state_col, "state")
    dept_col = detect_col(headers, args.dept_col, "dept")
    desig_col = detect_col(headers, args.designation_col, "designation")
    statename_col = detect_col(headers, "", "statename")
    deptname_col = detect_col(headers, "", "deptname")
    orgtype_col = detect_col(headers, "", "orgtype")
    log.info("detected columns -> " + ", ".join(f"{k}={v!r}" for k, v in [
        ("state", state_col), ("dept", dept_col), ("designation", desig_col), ("org_type", orgtype_col),
        ("state_name", statename_col), ("dept_name", deptname_col)]))

    # All six columns are mandatory (state_center_id, department_id, org_type,
    # state_center_name, department_name, designation) -- no CLI fallback for any of them.
    missing = [label for col, label in [
        (state_col, "state_center_id"), (dept_col, "department_id"), (orgtype_col, "org_type"),
        (statename_col, "state_center_name"), (deptname_col, "department_name"), (desig_col, "designation"),
    ] if not col]
    if missing:
        raise SystemExit(f"[config] source file is missing required column(s): {missing}. "
                         f"Headers found: {headers}")

    # No filter column in this file format -- every row is processed.
    filtered = rows
    if not filtered:
        raise SystemExit("[config] source file has no data rows.")

    # id numeric-corruption guard
    for col, label in [(state_col, "state"), (dept_col, "dept")]:
        vals = [_s(r.get(col)) for r in filtered if _s(r.get(col))]
        bad = [v for v in vals if re.search(r"[eE]\+?\d|\.\d", v)]
        if bad:
            log.warning(f"!! {label} id column {col!r}: {len(bad)}/{len(vals)} values look NUMERIC "
                        f"(e.g. {bad[0]!r}) -- long ids likely CORRUPTED by Excel; matches will fail. "
                        f"Format the column as TEXT or supply a CSV.")

    cols = {"state": state_col, "dept": dept_col, "designation": desig_col,
            "statename": statename_col, "deptname": deptname_col, "orgtype": orgtype_col}
    targets = await resolve_targets(filtered, cols, user_uuid)

    log.info(f"resolution: {_tally(targets, 'resolution')}")
    matched = [t for t in targets if t.resolution == "matched"]
    log.info(f"matched by prior status: {_tally(matched, 'prior_status')}")

    to_process = [t for t in matched if t.prior_status == "NOT_STARTED" or (args.force and t.prior_status == "COMPLETED")]
    if args.limit:
        to_process = to_process[:args.limit]
    log.info(f"TO PROCESS: {len(to_process)} designations  (force={args.force}, limit={args.limit or 'none'})")

    annotator = None
    if not args.no_annotate:
        is_csv = args.excel.lower().endswith(".csv")
        status_out = args.status_out or (args.excel if is_csv
                                         else str(Path(args.excel).with_suffix("")) + ".rolemap_status.csv")
        annotator = SourceAnnotator(status_out, headers, rows, targets, cols, include_origin=not is_csv)
        log.info(f"per-row status -> {status_out}")

    out = args.out or str(Path(args.excel).with_name(
        f"role_mapping_{'run' if args.execute else 'plan'}_"
        f"{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"))

    if not args.execute:
        write_report(out, targets)
        if annotator is not None:
            annotator.write()
            log.info(f"per-row status written: {annotator.path}")
        log.info("DRY RUN complete -- no role mappings generated. Re-run with --execute to process.")
        await sessionmanager.close()
        return

    report = Report(out)
    proc_ids = {id(t) for t in to_process}
    for t in targets:
        if id(t) not in proc_ids:
            t.final_status = t.final_status or t.prior_status
            report.add_sync(t)
    if annotator is not None:
        annotator.write()
    log.info(f"EXECUTING {len(to_process)} role mappings (PASS 2 + PASS 4), {args.batch_size} at a time "
             f"(retries={args.retries}, per-doc timeout={PER_DOC_TIMEOUT}s) ...")
    try:
        await run_execute(to_process, user_uuid, args.instruction or None, args.force,
                          args.igot_match, args.batch_size, args.retries, report, annotator)
    finally:
        report.close()
        if annotator is not None:
            annotator.write()
    log.info(f"final status: {_tally(to_process, 'final_status')}")
    ttot = sum(t.tok_total for t in to_process)
    log.info(f"TOTAL tokens over {len(to_process)} designations: "
             f"input={sum(t.tok_input for t in to_process)} output={sum(t.tok_output for t in to_process)} "
             f"thinking={sum(t.tok_thinking for t in to_process)} total={ttot}")
    log.info(f"report: {out}")
    if annotator is not None:
        log.info(f"per-row status: {annotator.path}")
    await sessionmanager.close()


if __name__ == "__main__":
    asyncio.run(main())